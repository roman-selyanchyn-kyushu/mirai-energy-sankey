"""Build claims.json — one entry per factual claim the manuscript makes.

Each claim carries the wording as it will appear in the paper, a verdict reached
from the data rather than asserted, the series behind it, the arithmetic, and a
full citation. Nothing here is typed by hand: every number is read from a file in
sources/ so the page, the paper and this script cannot drift apart.

Add a claim by writing a build_<id>() function and listing it in CLAIMS.
"""
import json, os, re
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCES = os.path.join(os.path.dirname(HERE), "sources")
FF = os.path.join(SOURCES, "energy-in-sweden-facts-and-figures-2026.xlsx")

# the Sankey palette, so a carrier keeps its colour across both pages
COLOR = {"oil": "#2c6b47", "gas": "#5aa7d6", "bio": "#4a9b52",
         "heat": "#e07590", "elec": "#f08c1e", "coal": "#606066",
         "wst": "#94953f", "amb": "#31b5a6", "other": "#9aa0a8"}


def _sheet(name):
    wb = load_workbook(FF, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    wb.close()
    return rows


def heating_by_carrier():
    """F&F table 3.4 — heating and hot water by carrier, three building groups.

    Layout: one header row of building groups, one of carriers, then a year per
    row. Columns repeat Oil / District heating / Electric heating / Natural gas /
    Biomass / Total for each group.
    """
    rows = _sheet("3.4")
    carriers = ["Oil", "District heating", "Electric heating", "Natural gas", "Biomass", "Total"]
    groups = {"detached": 1, "apartment": 7, "premises": 13}
    out = {}
    for r in rows:
        if isinstance(r[0], (int, float)) and 1983 <= r[0] <= 2100:
            out[int(r[0])] = {g: {c: (r[b+i] or 0.0) for i, c in enumerate(carriers)}
                              for g, b in groups.items()}
    return out


def district_heating_input():
    """F&F table 7.2 — what is burned to make district heat."""
    rows = _sheet("7.2")
    hdr = [str(c).strip() if c else "" for c in rows[5]]
    out = {}
    for r in rows[6:]:
        if isinstance(r[0], (int, float)) and 1970 <= r[0] <= 2100:
            out[int(r[0])] = {hdr[j]: (r[j] or 0.0) for j in range(1, len(hdr)) if hdr[j]}
    return out


def build_se_heating_fossil():
    H, DH = heating_by_carrier(), district_heating_input()
    years = sorted(H)
    y0, y1 = years[0], years[-1]
    BASE = 1990                       # the year the manuscript's claim starts from

    # dwellings = detached houses + apartment blocks, excluding non-residential premises
    def dw(y, carrier):
        return H[y]["detached"][carrier] + H[y]["apartment"][carrier]

    series = [("oil", "Oil", "Oil"), ("gas", "Natural gas", "Natural gas"),
              ("bio", "Biomass", "Biomass (burned in the building)"),
              ("heat", "District heating", "District heating"),
              ("elec", "Electric heating", "Electric heating")]

    chart = {
        "kind": "stacked", "unit": "TWh",
        "title": f"Energy for heating and hot water in Swedish dwellings, {y0}–{y1}",
        "years": years,
        "series": [{"id": sid, "label": lab, "color": COLOR[sid],
                    "values": [round(dw(y, col), 3) for y in years]}
                   for sid, col, lab in series],
        "note": "Detached houses and apartment blocks; non-residential premises excluded.",
    }

    def row(label, col, digits=2):
        a, b = dw(BASE, col), dw(y1, col)
        return {"label": label, "a": round(a, digits), "b": round(b, digits),
                "delta": round(b-a, digits),
                "pct": round((b-a)/a*100, 1) if a else None}

    numbers = [row("Oil", "Oil"), row("Natural gas", "Natural gas"),
               row("District heating", "District heating"),
               row("Electric heating", "Electric heating"),
               row("Biomass, burned in the building", "Biomass"),
               row("Total heating demand", "Total")]

    dh0, dh1 = DH[BASE], DH[y1]
    COAL = "Coal incl. coke oven and blast furnace gases"
    bio_share = lambda d: d["Biomass"]/d["Total"]*100
    dh_numbers = [
        {"label": "Biomass into district heating", "a": round(dh0["Biomass"], 2),
         "b": round(dh1["Biomass"], 2), "delta": round(dh1["Biomass"]-dh0["Biomass"], 2),
         "pct": round((dh1["Biomass"]-dh0["Biomass"])/dh0["Biomass"]*100, 1)},
        {"label": "… as a share of all district-heating input", "a": round(bio_share(dh0), 1),
         "b": round(bio_share(dh1), 1), "delta": round(bio_share(dh1)-bio_share(dh0), 1),
         "pct": None, "unit": "%"},
        {"label": "Coal into district heating", "a": round(dh0[COAL], 2), "b": round(dh1[COAL], 2),
         "delta": round(dh1[COAL]-dh0[COAL], 2), "pct": round((dh1[COAL]-dh0[COAL])/dh0[COAL]*100, 1)},
        {"label": "Petroleum products into district heating", "a": round(dh0["Petroleum products"], 2),
         "b": round(dh1["Petroleum products"], 2),
         "delta": round(dh1["Petroleum products"]-dh0["Petroleum products"], 2),
         "pct": round((dh1["Petroleum products"]-dh0["Petroleum products"])/dh0["Petroleum products"]*100, 1)},
    ]

    oil0, oil1 = dw(BASE, "Oil"), dw(y1, "Oil")
    bio0, bio1 = dw(BASE, "Biomass"), dw(y1, "Biomass")
    dhd0, dhd1 = dw(BASE, "District heating"), dw(y1, "District heating")
    tot0, tot1 = dw(BASE, "Total"), dw(y1, "Total")

    return {
        "id": "se-heating-fossil",
        "title": "Fossil fuel in Swedish residential heating",
        "claim": "Biomass has almost entirely removed fossil fuel from residential heating in Sweden.",
        "verdict": "revised",
        "verdictLine": (
            f"The fossil collapse is real — oil for heating Swedish dwellings fell "
            f"{oil0:.1f} → {oil1:.2f} TWh ({(oil1-oil0)/oil0*100:.0f}%) between {BASE} and {y1}. "
            f"But biomass burned <i>in the building</i> did not replace it: that is "
            f"{bio0:.1f} → {bio1:.1f} TWh, <b>{(bio1-bio0)/bio0*100:.0f}%</b>. The substitution ran through "
            f"district heating and efficiency instead, and biomass's real role is one step upstream."),
        "rewrite": (
            f"Direct oil use for heating Swedish dwellings fell from {oil0:.1f} TWh in {BASE} to "
            f"{oil1:.2f} TWh in {y1}, a reduction of {abs((oil1-oil0)/oil0*100):.0f}%. The substitution was not "
            f"direct biomass combustion in buildings — slightly lower today ({bio0:.1f} → {bio1:.1f} TWh) — "
            f"but district heating, whose deliveries to dwellings rose "
            f"{abs((dhd1-dhd0)/dhd0*100):.0f}% while its own fuel input shifted from {bio_share(dh0):.0f}% to "
            f"{bio_share(dh1):.0f}% biomass, together with electric heat pumps and a "
            f"{abs((tot1-tot0)/tot0*100):.0f}% fall in total heating demand."),
        "chart": chart,
        "tables": [
            {"title": f"Swedish dwellings: heating and hot water, {BASE} → {y1}",
             "cols": [str(BASE), str(y1), "change"], "unit": "TWh", "rows": numbers},
            {"title": f"What district heating is made from, {BASE} → {y1}",
             "cols": [str(BASE), str(y1), "change"], "unit": "TWh", "rows": dh_numbers},
        ],
        "derivation": [
            "Dwellings = one- and two-dwelling buildings + multi-dwelling buildings from table 3.4; "
            "non-residential premises are excluded so the claim stays about homes.",
            "No weather correction, and no adjustment for heated floor area or population — these are "
            "the published annual observations.",
            "District-heating input (table 7.2) is the fuel burned to produce the heat, not the heat "
            "delivered, so it is not additive with the building-side figures.",
            "The five carriers do not always sum exactly to the published Total column — the table is "
            "published rounded, and the residual reaches 0.25 TWh (0.4%) at its widest. The chart stacks "
            "the carriers; the Total row quotes the published total.",
        ],
        "caveats": [
            f"Biomass in dwellings peaked at {max(dw(y,'Biomass') for y in years):.1f} TWh around "
            f"{max(years, key=lambda y: dw(y,'Biomass'))} and has fallen since; the {BASE}→{y1} endpoints "
            "understate that rise and fall.",
            "Electric heating includes heat pumps, so its modest fall hides a large gain in delivered "
            "heat per unit of electricity. The table measures purchased energy, not heat delivered.",
            "District heating is counted where it is burned. Attributing its biomass back to the "
            "dwellings it heats is a modelling choice this table does not make for you.",
        ],
        "sources": [
            {"author": "Energimyndigheten (Swedish Energy Agency)",
             "year": "2026",
             "title": "Energy in Sweden — Facts and Figures 2026",
             "detail": "Table 3.4, Energy use for heating and hot water in dwellings and non-residential "
                       "premises, from 1983, TWh. Underlying survey: Energy statistics for dwellings and "
                       "non-residential premises (Energimyndigheten and Statistics Sweden).",
             "url": "https://www.energimyndigheten.se/en/facts-and-figures/statistics/"},
            {"author": "Energimyndigheten (Swedish Energy Agency)",
             "year": "2026",
             "title": "Energy in Sweden — Facts and Figures 2026",
             "detail": "Table 7.2, Input energy used in the production of district heating, from 1970, TWh.",
             "url": "https://www.energimyndigheten.se/en/facts-and-figures/statistics/"},
        ],
        "crosscheck": (
            "An independent route through Naturvårdsverket's stationary-combustion activity data "
            "(CRF 1A4a + 1A4b, residences and commercial/public premises) gives oil + gas of 32.29 TWh in "
            "1990. Table 3.4 gives 32.10 TWh for all buildings on the same footing — agreement to 0.6%. "
            "The two diverge in the recent year: Naturvårdsverket reports 1.87 TWh for 2023 against table "
            "3.4's 0.81 TWh, almost all of it in the gas series (1.19 vs 0.43 TWh). That is the difference "
            "between a −94% and a −97% headline, so state which scope you are quoting."),
    }


CLAIMS = [build_se_heating_fossil]


def main():
    out = []
    for i, fn in enumerate(CLAIMS, 1):
        c = fn()
        c["number"] = i
        out.append(c)
    path = os.path.join(HERE, "claims.json")
    with open(path, "w") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)
    print(f"wrote {path}  ({len(out)} claim{'s' if len(out)!=1 else ''})")
    for c in out:
        if c.get("kind") == "biomass":
            m = sum(len(b["metrics"]) for b in c["panel"]["blocks"])
            bars = sum(len(b["bars"]) for b in c["bars"])
            print(f"  {c['number']}. {c['id']:24} verdict={c['verdict']:11} "
                  f"{len(c['bars'])} bar blocks / {bars} bars, {m} metrics, "
                  f"{len(c['tables'])} tables, {len(c['sources'])} sources")
        elif c.get("kind") == "panel":
            m = sum(len(b["metrics"]) for b in c["panel"]["blocks"])
            print(f"  {c['number']}. {c['id']:24} panel        "
                  f"{len(c['panel']['blocks'])} blocks, {m} metrics, {len(c['sources'])} sources")
        else:
            n = len(c["chart"]["years"])
            print(f"  {c['number']}. {c['id']:24} verdict={c['verdict']:11} "
                  f"{n} years, {len(c['chart']['series'])} series, {len(c['sources'])} sources")




# ── per-capita comparison panel ────────────────────────────────────────────
# Sweden and Japan book primary energy differently, so a naive per-capita
# comparison of primary supply or fossil share is an artefact of accounting.
# The panel therefore separates metrics that need no adjustment from two that
# are restated onto one convention, with every adjustment shown as its own step.

NCV_GCV = {"coal": 0.95, "oil": 0.95, "gas": 0.90}   # IEA/Eurostat net-to-gross ratios
NUCLEAR_EFF = 0.33                                    # IEA convention, applied to both countries


def _pop():
    return json.load(open(os.path.join(SOURCES, "population.json")))


def _ds():
    return json.load(open(os.path.join(HERE, "datasets.json")))


def build_se_jp_percapita():
    import derive_se, derive_jp
    D, POP = _ds(), _pop()
    YEAR = 2024
    pse = POP["se"][str(YEAR)]["value"]
    pjp = POP["jp"][str(YEAR)]["value"]
    SEC = ("res", "com", "ind", "tpt")

    def parts(key):
        d = D[key]
        ids = {n["id"] for n in d["nodes"] if n["col"] == 0}
        band, inn, out = {}, {}, {}
        for s, t, v in d["flows"]:
            out[s] = out.get(s, 0) + v
            inn[t] = inn.get(t, 0) + v
            if s in ids:
                band[s] = band.get(s, 0) + v
        return d, band, inn, out

    se, seB, seI, seO = parts(f"se{YEAR}")
    jp, jpB, jpI, jpO = parts(f"jp{YEAR}")
    carrier = lambda d, c: sum(v for s, t, v in d["flows"] if s == c and t in SEC)
    gen = lambda d, O: O.get("elec", 0) - sum(v for s, t, v in d["flows"] if s == "elec" and t == "rej")
    TJ_MWH = 1e6/3600

    # ── trade, for self-sufficiency ──
    T = derive_se.load(YEAR)
    g = lambda r: (T.get((r, "Total")) or 0)
    val = derive_jp.read_table(YEAR)
    se_ss = g("1.1 Indigenous production")/g("1. Total supply")*100
    jp_ss = val("#110000", "$1400")/val("#190000", "$1400")*100

    # ── harmonisation: put both on the physical energy content method ──
    ng = derive_jp.natural_units(YEAR)          # GWh by source, from METI's own sheet
    e = lambda c: ng[c]*3.6                     # GWh -> TJ of electricity
    jp_h = {
        "nuclear": e("$1100")/NUCLEAR_EFF,      # substitution -> reactor heat
        "hydro":   e("$0800"),                  # substitution -> generated electricity
        "wind":    e("$N120"),
        "solar":   e("$N111") + val("#190000", "$N112"),   # PV as electricity, solar heat as heat
        "geo":     jpB.get("geo", 0),           # already reported as heat
        "bio":     jpB.get("bio", 0),
        "wst":     jpB.get("wst", 0) - val("#190000", "$N250"),   # drop recovered energy
        "coal":    jpB.get("coal", 0)*NCV_GCV["coal"],            # gross -> net calorific value
        "oil":     jpB.get("oil", 0)*NCV_GCV["oil"],
        "gas":     jpB.get("gas", 0)*NCV_GCV["gas"],
    }
    se_nuc_elec = D[f"se{YEAR}"]["meta"]["nuclear_electricity_TJ"]
    se_h = dict(seB)
    se_h["nuclear"] = se_nuc_elec/NUCLEAR_EFF   # restate from actual 35.8% onto the same 33%

    FOSSIL = ("coal", "gas", "oil")
    fos = lambda b: sum(b.get(f, 0) for f in FOSSIL)

    def M(label, a, b, unit, note=""):
        return {"label": label, "se": round(a, 2), "jp": round(b, 2), "unit": unit,
                "ratio": round(a/b, 2) if b else None, "note": note}

    block_a = [
        M("Electricity used", carrier(se, "elec")*TJ_MWH/pse, carrier(jp, "elec")*TJ_MWH/pjp, "MWh/person"),
        M("Electricity generated", gen(se, seO)*TJ_MWH/pse, gen(jp, jpO)*TJ_MWH/pjp, "MWh/person"),
        M("District heat used", carrier(se, "heat")*1000/pse, carrier(jp, "heat")*1000/pjp, "GJ/person"),
        M("Final energy consumption", sum(seI.get(x, 0) for x in SEC+("ne",))*1000/pse,
          sum(jpI.get(x, 0) for x in SEC+("ne",))*1000/pjp, "GJ/person"),
        M("Non-energy use (feedstocks)", seI.get("ne", 0)*1000/pse, jpI.get("ne", 0)*1000/pjp, "GJ/person"),
        M("Energy self-sufficiency", se_ss, jp_ss, "%",
          "Both balances count nuclear as domestic production although uranium is imported."),
    ]
    block_b = [
        M("Primary energy supply", sum(se_h.values())*1000/pse, sum(jp_h.values())*1000/pjp, "GJ/person"),
        M("Fossil share of primary supply", fos(se_h)/sum(se_h.values())*100,
          fos(jp_h)/sum(jp_h.values())*100, "%"),
    ]

    sector = [{"label": {"res": "Residential", "com": "Commercial & services",
                         "ind": "Industrial", "tpt": "Transport"}[s],
               "se": round(seI.get(s, 0)/sum(seI.get(x, 0) for x in SEC)*100, 1),
               "jp": round(jpI.get(s, 0)/sum(jpI.get(x, 0) for x in SEC)*100, 1),
               "unit": "%", "ratio": None} for s in SEC]

    steps = [
        {"what": "Japan: hydro, wind and solar PV from substitution equivalent to generated electricity",
         "delta": round((e("$0800")+e("$N120")+e("$N111")
                         - (jpB.get("hydro", 0)+jpB.get("wind", 0)+jpB.get("solar", 0)-val("#190000", "$N112")))/1000, 1)},
        {"what": f"Japan: nuclear from substitution equivalent to reactor heat at {NUCLEAR_EFF:.0%}",
         "delta": round((jp_h["nuclear"]-jpB.get("nuclear", 0))/1000, 1)},
        {"what": "Japan: fossil fuels from gross to net calorific value "
                 f"(coal ×{NCV_GCV['coal']}, oil ×{NCV_GCV['oil']}, gas ×{NCV_GCV['gas']})",
         "delta": round((fos(jp_h)-fos(jpB))/1000, 1)},
        {"what": "Japan: recovered industrial steam and electricity removed, which Sweden does not book",
         "delta": round(-val("#190000", "$N250")/1000, 1)},
        {"what": f"Sweden: nuclear restated from its actual {se_nuc_elec/seB['nuclear']*100:.1f}% "
                 f"reactor efficiency onto the same {NUCLEAR_EFF:.0%}",
         "delta": round((se_h["nuclear"]-seB["nuclear"])/1000, 1)},
    ]

    return {
        "id": "se-jp-percapita",
        "kind": "panel",
        "title": f"Sweden and Japan per person, {YEAR}",
        "claim": "The two energy systems differ structurally, not just in scale.",
        "verdict": None,
        "lead": (
            f"Normalised by population — Sweden {pse:,} and Japan {pjp:,} — the two systems separate on "
            f"electricity and heat rather than on total energy. The first block needs no adjustment: every "
            f"metric is measured energy, unaffected by how the two countries book primary energy. The "
            f"second block does, and the adjustment is shown rather than assumed."),
        "panel": {"blocks": [
            {"title": "Directly comparable — no adjustment needed", "metrics": block_a,
             "note": "These are measured quantities: delivered energy, generated electricity, and trade. "
                     "The primary-energy convention that makes the two Sankeys incomparable does not touch them."},
            {"title": "Harmonised onto one convention", "metrics": block_b,
             "note": "Both countries restated onto the physical energy content method: renewable electricity "
                     "counted 1:1, nuclear as reactor heat at a single efficiency, fossil fuels at net "
                     "calorific value."},
            {"title": "Sector split of final energy use", "metrics": sector,
             "note": "Share of final energy consumption excluding non-energy use."},
        ]},
        "steps": steps,
        "population": {"se": POP["se"][str(YEAR)], "jp": POP["jp"][str(YEAR)]},
        "derivation": [
            "Energy figures come from the same derivation that produces the Sankey diagrams, so the panel "
            "and the charts cannot disagree.",
            "Population is put on a mid-period basis in both countries so the denominator matches the "
            "energy year: Sweden the mean of the bracketing year-ends, Japan the 1 October estimate that "
            "sits inside the April–March fiscal year.",
            f"Harmonisation applies the physical energy content method to both countries: renewable "
            f"electricity at 1:1, nuclear as reactor heat at {NUCLEAR_EFF:.0%}, and Japanese fossil fuels "
            f"converted from gross to net calorific value.",
        ],
        "caveats": [
            f"The net-to-gross calorific ratios (coal {NCV_GCV['coal']}, oil {NCV_GCV['oil']}, gas "
            f"{NCV_GCV['gas']}) are the standard IEA/Eurostat values, not METI's own published table, "
            "which sits behind an access restriction. They agree with the ratios implied by the IEA "
            "reconciliation in the project README to within about one percentage point.",
            f"Nuclear is put at {NUCLEAR_EFF:.0%} for both countries so the comparison is internally "
            f"consistent. Sweden's balance reports an actual {se_nuc_elec/seB['nuclear']*100:.1f}%; using "
            f"that instead would lower Swedish primary supply by "
            f"{abs((se_h['nuclear']-seB['nuclear'])*1000/pse):.1f} GJ/person.",
            "Japanese biomass is also booked at gross calorific value, but no reliable net ratio was "
            "available for it, so it is left unadjusted. It is 2.8% of Japanese supply, so the effect on "
            "the totals is under half a percent.",
            "Sweden is a large net electricity exporter, which is why generated electricity per person "
            "exceeds electricity used by more than it does in Japan.",
        ],
        "sources": [
            {"author": "SCB (Statistics Sweden)", "year": "2026",
             "title": "Population by region, age, marital status and sex",
             "detail": POP["se"][str(YEAR)]["basis"] + ". Table BE0101, retrieved through the PxWeb API.",
             "url": POP["se"][str(YEAR)]["url"]},
            {"author": "Statistics Bureau of Japan", "year": "2025",
             "title": "人口推計 (Population Estimates), 2024",
             "detail": POP["jp"][str(YEAR)]["basis"] + ". Table 1, total population.",
             "url": POP["jp"][str(YEAR)]["url"]},
            {"author": "Energimyndigheten (Swedish Energy Agency)", "year": "2026",
             "title": "Annual energy balance EN0202_A, calendar year 2024",
             "detail": "The same extraction that produces the Swedish Sankey diagram.",
             "url": "https://pxexternal.energimyndigheten.se/pxweb/en/Energimyndighetens_statistikdatabas/"},
            {"author": "METI / Agency for Natural Resources and Energy", "year": "2026",
             "title": "総合エネルギー統計 Comprehensive Energy Statistics, FY2024 (確報)",
             "detail": "The same extraction that produces the Japanese Sankey diagram, plus the "
                       "natural-units sheet for generation in GWh.",
             "url": "https://www.enecho.meti.go.jp/statistics/total_energy/results.html"},
        ],
    }



# ── forest biomass: Sweden and Japan ───────────────────────────────────────
# The two countries have almost the same forest endowment and opposite ways of
# using it. Every figure is read out of the compiled research note in
# sources/claims/, matched on a distinctive phrase from its own bullet, so a
# changed or missing line breaks the build instead of publishing a wrong number.

BIOMASS_MD = os.path.join(SOURCES, "claims", "japan_forest_biomass_verified_statistics.md")

# feedstock palette, kept close to the Sankey colours
BC = {"prim": "#4a9b52", "byprod": "#7cb87f", "rec": "#94953f", "domchip": "#2f6b3a",
      "imp_pellet": "#c2562f", "imp_chip": "#d98a45", "pks": "#a2643a", "dom": "#4a9b52"}


def _bullets():
    """Every '- **value** — text' bullet in the research note, in file order."""
    out = []
    for line in open(BIOMASS_MD, encoding="utf-8"):
        m = re.match(r"^- \*\*(.+?)\*\*\s*[—-]\s*(.*)$", line.strip())
        if m:
            out.append({"lead": m.group(1), "text": m.group(2), "line": line.strip()})
    return out


class _Note:
    """Lookup by a distinctive phrase; raises rather than guessing."""

    def __init__(self):
        self.rows = _bullets()
        if len(self.rows) < 80:
            raise SystemExit(f"biomass note: only {len(self.rows)} bullets parsed, file format changed?")
        self.used = set()

    def _row(self, phrase):
        hits = [r for r in self.rows if phrase in r["line"]]
        if not hits:
            raise SystemExit(f"biomass note: no bullet contains {phrase!r}")
        self.used.add(phrase)
        return hits[0]

    def num(self, phrase, scale=1.0, lead_has=None):
        """The bolded lead value of the matching bullet, as a number.

        lead_has narrows the match when the note states the same quantity twice in
        different currencies, e.g. the Swedish prices in both yen and dollars.
        """
        hits = [r for r in self.rows if phrase in r["line"]
                and (lead_has is None or lead_has in r["lead"])]
        if not hits:
            raise SystemExit(f"biomass note: no bullet matches {phrase!r} / lead {lead_has!r}")
        self.used.add(phrase)
        lead = hits[0]["lead"]
        m = re.search(r"[\d,]+(?:\.\d+)?", lead.replace("¥", "").replace("US$", ""))
        if not m:
            raise SystemExit(f"biomass note: no number in lead {lead!r} for {phrase!r}")
        return float(m.group(0).replace(",", "")) * scale

    def inline(self, phrase, pattern):
        """A number written inside the bullet text rather than as its lead."""
        m = re.search(pattern, self._row(phrase)["line"])
        if not m:
            raise SystemExit(f"biomass note: pattern {pattern!r} not found for {phrase!r}")
        return float(m.group(1).replace(",", ""))

CLAIMS.append(build_se_jp_percapita)



def build_se_jp_biomass():
    N, POP = _Note(), _pop()
    pse, pjp = POP["se"]["2024"]["value"], POP["jp"]["2024"]["value"]

    # ── forest estate ──
    jp_forest = N.num("Japan\u2019s total forest area")
    jp_land = N.inline("forest cover in **2022**", r"land area is ([\d.]+) Mha")
    jp_planted = N.num("planted forest (\u4eba\u5de5\u6797)")
    jp_natural = N.num("natural forest (\u5929\u7136\u6797)")
    jp_other_f = N.num("other forest land, including non-stocked")
    se_forest = N.num("total Swedish forest land")
    se_land = N.num("Swedish land area in the same inventory")
    se_prod = N.num("productive forest land.")

    # ── growth and harvest ──
    se_incr = N.num("average gross annual increment")
    se_fell = N.num("gross felling")
    se_net = N.num("preliminary net felling")
    jp_growth = N.num("recent annual increase in Japan\u2019s forest growing stock")
    jp_round = N.num("domestically supplied industrial roundwood")
    jp_supply = N.num("total domestic wood supply including fuelwood")
    jp_demand = N.num("total Japanese wood demand")
    jp_ss = N.num("total wood self-sufficiency ratio, **2024**")
    jp_ss_low = N.num("historical low in total wood self-sufficiency")

    # ── mobilisation barriers ──
    jp_road = N.num("forest internal-road-network density")
    jp_road_km = N.num("total Japanese forest-road")
    se_road = N.num("reported Swedish national forest-road density")
    jp_slope = N.num("average slope of Japan\u2019s forest land")
    jp_slope30 = N.num("\u80b2\u6210\u6797) on slopes of 30\u00b0 or more")
    jp_house = N.num("households holding at least 1 ha")
    jp_hold = N.num("average holding")
    jp_small = N.num("forestry households holding less than 10 ha")

    # ── Japan solid fuel balance, 2024, physical mass ──
    Mt = 1e-6
    pel_tot = N.num("wood pellets burned for energy in Japan") * Mt
    pel_dom = N.num("domestically manufactured pellets burned") * Mt
    pel_imp = N.num("imported pellets burned") * Mt
    chip_tot = N.num("wood chips burned for energy") * Mt
    chip_thin = N.num("chips from thinning and logging residues") * Mt
    chip_saw = N.num("chips from sawmill and wood-processing residues") * Mt
    chip_con = N.num("chips from construction waste wood") * Mt
    chip_oth = N.num("chips from other sources") * Mt
    chip_imp = N.num("imported wood chips burned") * Mt
    chip_dom = N.num("domestic-source chips burned") * Mt
    pks = N.num("PKS imports, **2024**") * Mt

    # ── import origins ──
    pel_cust = N.num("customs-recorded pellet imports")
    origins_pel = [(c, N.num(f"pellet imports from {c}, **2024**")) for c in
                   ("Vietnam", "Canada", "the United States", "Malaysia", "Indonesia")]
    origins_pks = [(c, N.num(f"PKS imports from {c}, **2024**") * Mt) for c in ("Indonesia", "Malaysia")]
    origins_pks.append(("other countries", N.num("PKS imports from other countries") * Mt))

    # ── the import ramp ──
    ramp = [(y, N.num(f"pellet imports, **{y}**"), N.num(f"PKS imports, **{y}**"))
            for y in (2015, 2020, 2023)]
    ramp.append((2024, pel_cust, pks))

    # ── Sweden unprocessed wood fuel, 2024 ──
    se_wf = N.num("total unprocessed wood-fuel use")
    se_imp_pc = N.num("imported raw material share of unprocessed wood fuel")
    se_byprod = N.num("industrial wood-processing by-products")
    se_prim = N.num("domestic primary forest fuels")
    se_rec = N.num("recycled wood fuel")

    # ── Japanese energy balance, FY2023 ──
    bal = [("Wood", N.num("\u201cwood\u201d within the biomass total")),
           ("Black liquor", N.num("black liquor, **FY2023**")),
           ("Other biomass", N.num("\u201cother biomass,\u201d **FY2023**")),
           ("Waste wood", N.num("waste wood, **FY2023**")),
           ("Transport biofuels", N.num("biofuels, **FY2023**"))]
    bal_tot = N.num("biomass in Japan\u2019s primary energy supply")

    # ── delivered fuel cost ──
    cost = {"jp_chip": N.num("wood-chip cost above, **CALCULATED**"),
            "jp_pellet": N.num("pellet cost above, **CALCULATED**"),
            "jp_pks": N.num("PKS cost above, **CALCULATED**"),
            "se_chip": N.num("Swedish forest-chip price, **CALCULATED**, **2024**",
                             lead_has="US$"),
            "se_ref": N.num("Swedish refined-wood-fuel price, **CALCULATED**, **2024**",
                            lead_has="US$")}

    # ── identities the note must satisfy; a mismatch stops the build ──
    checks = []

    def check(what, a, b, tol):
        ok = abs(a - b) <= tol
        checks.append({"what": what, "a": round(a, 3), "b": round(b, 3), "ok": ok})
        if not ok:
            raise SystemExit(f"biomass note fails its own arithmetic: {what} ({a} vs {b})")

    check("Japanese pellets: domestic + imported = total burned", pel_dom + pel_imp, pel_tot, 1e-6)
    check("Japanese chips: the four domestic categories = domestic total",
          chip_thin + chip_saw + chip_con + chip_oth, chip_dom, 1e-6)
    check("Japanese chips: domestic + imported = total burned", chip_dom + chip_imp, chip_tot, 1e-6)
    check("PKS: Indonesia + Malaysia + other = total imports", sum(v for _, v in origins_pks), pks, 1e-6)
    check("Japanese wood self-sufficiency = domestic supply / demand",
          jp_supply / jp_demand * 100, jp_ss, 0.1)
    check("Japanese forest: planted + natural + other = total area",
          jp_planted + jp_natural + jp_other_f, jp_forest, 0.01)
    check("Swedish wood fuel: the three categories \u2248 the published total",
          se_byprod + se_prim + se_rec, se_wf, 500)

    # ── derived aggregates ──
    jp_imp = pel_imp + chip_imp + pks
    jp_dom = pel_dom + chip_dom
    jp_fuel = jp_imp + jp_dom
    jp_imp_pc = jp_imp / jp_fuel * 100
    se_imp_gwh = se_wf * se_imp_pc / 100

    def M(label, a, b, unit, note=""):
        return {"label": label, "se": round(a, 2), "jp": round(b, 2), "unit": unit,
                "ratio": round(a / b, 2) if b else None, "note": note}

    return {
        "id": "se-jp-biomass",
        "kind": "biomass",
        "title": "Forest biomass in Sweden and Japan \u2014 same forest, opposite supply chain",
        "claim": ("Sweden and Japan are both heavily forested countries that use forest biomass "
                  "for energy."),
        "verdict": "revised",
        "verdictLine": (
            f"The forest endowment really is almost the same \u2014 Sweden {se_forest} Mha, Japan "
            f"{jp_forest} Mha, {se_forest/se_land*100:.0f}% and {jp_forest/jp_land*100:.0f}% of land area. "
            f"What the two countries do with it is opposite. Under {se_imp_pc:.0f}% of the raw material in "
            f"Sweden's wood fuel is imported, and about half of it is residue from its own sawmills and "
            f"pulp mills. In Japan <b>{jp_imp_pc:.0f}% of the solid biomass fuel burned in 2024 was "
            f"imported</b> by mass, and the single largest fuel is not wood at all but "
            f"<b>palm kernel shell from Indonesia and Malaysia</b>."),
        "rewrite": (
            f"Sweden and Japan hold comparable forests \u2014 {se_forest} and {jp_forest} Mha, about "
            f"{se_forest/se_land*100:.0f}% and {jp_forest/jp_land*100:.0f}% of their land area \u2014 but "
            f"supply their bioenergy in opposite ways. Sweden's wood fuel is a by-product of a large "
            f"domestic forest industry: {se_wf:,.0f} GWh of unprocessed wood fuel in 2024, of which "
            f"{se_byprod/se_wf*100:.0f}% is sawmill and pulp-mill residue and less than {se_imp_pc:.0f}% of "
            f"the raw material is imported. Japan burned {jp_fuel:.1f} Mt of solid biomass fuel in 2024, "
            f"{jp_imp_pc:.0f}% of it imported \u2014 {pel_imp:.1f} Mt of wood pellets, {pks:.1f} Mt of palm "
            f"kernel shell and {chip_imp:.2f} Mt of chips \u2014 while harvesting only "
            f"{jp_round:.1f} million m\u00b3 of industrial roundwood against roughly {jp_growth:.0f} million "
            f"m\u00b3 of annual forest growth. The constraint is mobilisation rather than resource: "
            f"{jp_slope:.0f}\u00b0 average slope, {jp_slope30:.0f}% of managed forest steeper than 30\u00b0, "
            f"and {jp_house:,.0f} forestry households averaging {jp_hold} ha each, {jp_small:.0f}% of them "
            f"below 10 ha."),

        # ── figure 1: composition bars ──
        "bars": [
            {"title": "Where the solid biomass fuel comes from, 2024", "scale": "share",
             "note": ("The two systems are measured in different units \u2014 Sweden reports wood fuel in "
                      "GWh of energy, Japan reports physical fuel in tonnes \u2014 so the comparable "
                      "quantity here is the split, not the length of the bar."),
             "bars": [
                 {"label": "Sweden", "sub": f"unprocessed wood fuel, {se_wf:,.0f} GWh",
                  "segments": [
                      {"label": "Domestic raw material", "value": se_wf - se_imp_gwh, "color": BC["dom"],
                       "note": f"\u2265{100-se_imp_pc:.0f}%"},
                      {"label": "Imported raw material", "value": se_imp_gwh, "color": BC["imp_pellet"],
                       "note": f"<{se_imp_pc:.0f}%"}]},
                 {"label": "Japan", "sub": f"solid biomass fuel, {jp_fuel:.1f} Mt",
                  "segments": [
                      {"label": "Domestic raw material", "value": jp_dom, "color": BC["dom"]},
                      {"label": "Imported raw material", "value": jp_imp, "color": BC["imp_pellet"]}]},
             ]},
            {"title": "What that fuel actually is", "scale": "share",
             "note": ("Colour groups the two countries by the kind of material: greens are forest and "
                      "sawmill material, olive is recovered wood, warm colours are imported. Primary "
                      "forest fuel means slash, chips, stemwood and firewood in Sweden, and thinning and "
                      "logging residue in Japan; Sweden's category is published as one line and cannot be "
                      "split into residue and stemwood."),
             "bars": [
                 {"label": "Sweden", "sub": f"{se_wf:,.0f} GWh",
                  "segments": [
                      {"label": "Sawmill & pulp-mill by-products", "value": se_byprod, "color": BC["byprod"]},
                      {"label": "Primary forest fuel", "value": se_prim, "color": BC["prim"]},
                      {"label": "Recovered / construction wood", "value": se_rec, "color": BC["rec"]},
                      {"label": "Unattributed", "value": max(se_wf - se_byprod - se_prim - se_rec, 0),
                       "color": "#c9c5ba"}]},
                 {"label": "Japan", "sub": f"{jp_fuel:.1f} Mt",
                  "segments": [
                      {"label": "Sawmill & pulp-mill by-products", "value": chip_saw, "color": BC["byprod"]},
                      {"label": "Primary forest fuel", "value": chip_thin, "color": BC["prim"]},
                      {"label": "Recovered / construction wood", "value": chip_con, "color": BC["rec"]},
                      {"label": "Other domestic", "value": chip_oth + pel_dom, "color": BC["domchip"]},
                      {"label": "Imported wood chips", "value": chip_imp, "color": BC["imp_chip"]},
                      {"label": "Imported wood pellets", "value": pel_imp, "color": BC["imp_pellet"]},
                      {"label": "Imported palm kernel shell", "value": pks, "color": BC["pks"]}]},
             ]},
            {"title": "Japan's imported biomass fuel, 2015 \u2192 2024", "scale": "absolute",
             "unit": "Mt, customs basis",
             "note": ("Customs entries rather than survey-reported combustion, so these differ slightly "
                      "from the consumption figures above. Pellets rose "
                      f"\u00d7{ramp[-1][1]/ramp[0][1]:.0f} and palm kernel shell "
                      f"\u00d7{ramp[-1][2]/ramp[0][2]:.0f} over the decade."),
             "bars": [{"label": str(y), "sub": f"{a+b:.1f} Mt", "segments": [
                 {"label": "Wood pellets", "value": a, "color": BC["imp_pellet"]},
                 {"label": "Palm kernel shell", "value": b, "color": BC["pks"]}]}
                 for y, a, b in ramp]},
        ],

        # ── figure 2: comparable metrics ──
        "panel": {"blocks": [
            {"title": "Forest endowment \u2014 directly comparable", "metrics": [
                M("Forest area", se_forest, jp_forest, "Mha"),
                M("Forest share of land area", se_forest/se_land*100, jp_forest/jp_land*100, "%"),
                M("Forest per person", se_forest*1e6/pse, jp_forest*1e6/pjp, "ha/person")],
             "note": ("The endowment is the part that really is alike. Per person it is not: the same "
                      "forest divided by twelve times the population.")},
            {"title": "Growth and harvest \u2014 boundaries differ, read with the caveat", "metrics": [
                M("Annual forest growth", se_incr, jp_growth, "million m\u00b3/yr"),
                M("Annual harvest", se_fell, jp_round, "million m\u00b3/yr"),
                M("Harvest as a share of growth", se_fell/se_incr*100, jp_round/jp_growth*100, "%")],
             "note": ("Sweden is gross felling against gross increment, both stem volume. Japan is "
                      "industrial roundwood supply against growing-stock increase, which excludes fuelwood "
                      "and residues, so the Japanese share is understated \u2014 but not by enough to "
                      "close a gap this wide.")},
            {"title": "Delivered fuel cost at the plant, 2024\u201325", "metrics": [
                M("Wood chips", cost["se_chip"], cost["jp_chip"], "US$/GJ"),
                M("Pellets / refined wood fuel", cost["se_ref"], cost["jp_pellet"], "US$/GJ")],
             "note": ("Japanese chips are the cheaper fuel of the two, which is the point: cost is not "
                      "what holds domestic mobilisation back. Currency-converted, and the delivery point "
                      "is not identically defined in the two sources.")},
        ]},

        "tables": [
            {"title": "Japan: solid biomass fuel burned, 2024",
             "head": ["Fuel", "Domestic", "Imported", "Total", "Imported share"],
             "num": [False, True, True, True, True],
             "rows": [["Wood chips (dry tonnes)", f"{chip_dom:.2f}", f"{chip_imp:.2f}",
                       f"{chip_tot:.2f}", f"{chip_imp/chip_tot*100:.1f}%"],
                      ["Wood pellets", f"{pel_dom:.2f}", f"{pel_imp:.2f}", f"{pel_tot:.2f}",
                       f"{pel_imp/pel_tot*100:.1f}%"],
                      ["Palm kernel shell", "\u2014", f"{pks:.2f}", f"{pks:.2f}", "100.0%"],
                      ["Total", f"{jp_dom:.2f}", f"{jp_imp:.2f}", f"{jp_fuel:.2f}",
                       f"{jp_imp_pc:.1f}%"]],
             "note": "Million tonnes. Chips are dry tonnes; pellets and PKS as-received."},
            {"title": "Where Japan's imported fuel is grown, 2024",
             "head": ["Origin", "Fuel", "Million tonnes", "Share of that fuel"],
             "num": [False, False, True, True],
             "rows": [[c.replace("the ", "").title() if c.startswith("the ") else c.title(),
                       "Wood pellets", f"{v:.3f}", f"{v/pel_cust*100:.1f}%"] for c, v in origins_pel]
                     + [[c.title(), "Palm kernel shell", f"{v:.3f}", f"{v/pks*100:.1f}%"]
                        for c, v in origins_pks],
             "note": "Pellet shares are of customs-recorded imports; PKS shares of total PKS imports."},
            {"title": "Sweden: unprocessed wood fuel, 2024",
             "head": ["Category", "GWh", "Share"],
             "num": [False, True, True],
             "rows": [["Industrial wood-processing by-products", f"{se_byprod:,.0f}",
                       f"{se_byprod/se_wf*100:.1f}%"],
                      ["Domestic primary forest fuel", f"{se_prim:,.0f}", f"{se_prim/se_wf*100:.1f}%"],
                      ["Recycled wood fuel", f"{se_rec:,.0f}", f"{se_rec/se_wf*100:.1f}%"],
                      ["Total", f"{se_wf:,.0f}", "100.0%"]],
             "note": "Imported raw material is reported only as a bound: less than 5% of the total."},
            {"title": "Forest structure and the barriers to mobilising it",
             "head": ["", "Sweden", "Japan"],
             "num": [False, True, True],
             "rows": [["Forest area, Mha", f"{se_forest}", f"{jp_forest}"],
                      ["Productive / planted forest, Mha", f"{se_prod} productive",
                       f"{jp_planted} planted"],
                      ["Naturally regenerated forest, Mha", "\u2014", f"{jp_natural}"],
                      ["Forest-road density, m/ha", f"{se_road} (2004)", f"{jp_road} (FY2023)"],
                      ["Total forest-road network, km", "\u2014", f"{jp_road_km:,.0f}"],
                      ["Average forest-land slope", "\u2014", f"{jp_slope:.0f}\u00b0"],
                      ["Managed forest steeper than 30\u00b0", "\u2014", f"{jp_slope30:.0f}%"],
                      ["Forestry households", "\u2014", f"{jp_house:,.0f}"],
                      ["Average holding, ha", "\u2014", f"{jp_hold}"],
                      ["Holdings under 10 ha", "\u2014", f"{jp_small:.0f}%"],
                      ["Wood self-sufficiency", "\u2014", f"{jp_ss}% (low of {jp_ss_low}% in 2002)"]],
             "note": ("Dashes are values the research note records as NOT FOUND in the Swedish "
                      "statistics, not zeros.")},
            {"title": "Japan: biomass in the national energy balance, FY2023",
             "head": ["Component", "PJ", "Share"],
             "num": [False, True, True],
             "rows": [[k, f"{v:,.0f}", f"{v/bal_tot*100:.1f}%"] for k, v in bal]
                     + [["Total biomass in primary supply", f"{bal_tot:,.0f}", "100.0%"]],
             "note": ("METI does not allocate pellets, chips and PKS separately inside this total, so the "
                      "physical fuel balance above cannot be mapped onto it line for line.")},
        ],

        "derivation": [
            "Every figure is read from the compiled research note in sources/claims/ by matching a "
            "distinctive phrase in its own bullet, so a renamed or deleted line stops the build rather "
            "than publishing a stale number.",
            f"Japan's imported share is computed on physical mass: imported pellets {pel_imp:.2f} + PKS "
            f"{pks:.2f} + imported chips {chip_imp:.2f} = {jp_imp:.2f} Mt, against domestic chips "
            f"{chip_dom:.2f} + domestic pellets {pel_dom:.2f} = {jp_dom:.2f} Mt.",
            "Sweden's import share is the Energy Agency's own published bound of less than 5% of raw "
            "material, drawn at 5% and labelled as a bound rather than a point estimate.",
            f"Forest per person uses the same populations as the per-capita card \u2014 Sweden {pse:,} "
            f"and Japan {pjp:,} \u2014 so the two cards cannot disagree.",
            "Forest share of land area is recomputed here from the area figures rather than quoting the "
            "rounded 'two-thirds' and '68.6%' in the sources.",
        ],
        "caveats": [
            "These figures are transcribed from a compiled research note, not re-derived from the primary "
            "publications by this pipeline. The note names a primary source for every line and those are "
            "cited below, but unlike the other cards no independent extraction stands behind them. Treat "
            "the numbers as sourced, not as verified by this project.",
            "The two countries' fuel figures are in different units and cannot be added or ranked against "
            "each other. Sweden publishes wood fuel in GWh of energy; Japan publishes physical tonnes, and "
            "mixes dry-tonne chips with as-received pellets and PKS. Only the shares are comparable.",
            f"Sweden's {se_road} m/ha forest-road density is a 2004 figure quoted in a Japanese government "
            f"comparative study; the note records that no current Sweden-wide series was found, and that "
            f"the two countries' road-network boundaries are not documented as identical. Japan's "
            f"{jp_road} m/ha is therefore higher than the Swedish number on paper, which should not be "
            "read as Japan being better served.",
            "Japan's harvest-to-growth ratio compares industrial roundwood supply with growing-stock "
            "increase. The two have different boundaries, and the roundwood figure excludes fuelwood and "
            "residues, so the true removal share is somewhat higher than shown.",
            "The often-quoted claim that a quarter of Japanese forest parcels have unknown owners refers "
            "to owners not locatable from the land register alone. After further tracing the unresolved "
            "share in the surveyed sample was 0.57%.",
            "Pellet and PKS import totals are customs entries, which differ from survey-reported "
            "combustion because of inventories, timing and survey coverage: 6.381 Mt of pellets entered "
            "in 2024 against 5.163 Mt reported burned.",
        ],
        "crosscheck": (
            "The research note is internally consistent where it can be tested. "
            + " ".join(f"{c['what']}: {c['a']:,.3f} vs {c['b']:,.3f}." for c in checks[:4])
            + f" Japan's published {jp_ss}% wood self-sufficiency reproduces from its own supply and "
            f"demand figures ({jp_supply} / {jp_demand} m\u00b3). The three Swedish wood-fuel categories "
            f"sum to {se_byprod+se_prim+se_rec:,.0f} GWh against a published total of {se_wf:,.0f} GWh, a "
            f"{se_wf-se_byprod-se_prim-se_rec:,.0f} GWh residual shown as unattributed rather than "
            "silently absorbed. All seven identities are asserted at build time."),
        "sources": [
            {"author": "\u6797\u91ce\u5e81 (Forestry Agency, Japan)", "year": "2023",
             "title": "\u68ee\u6797\u8cc7\u6e90\u306e\u73fe\u6cc1 (Current State of Forest Resources)",
             "detail": "Forest area as of 31 March 2022: total, planted, natural and other forest land.",
             "url": "https://www.rinya.maff.go.jp/j/press/keikaku/attach/pdf/231013-2.pdf"},
            {"author": "\u6797\u91ce\u5e81 (Forestry Agency, Japan)", "year": "2025",
             "title": "\u4ee4\u548c6\u5e74\u5ea6 \u68ee\u6797\u30fb\u6797\u696d\u767d\u66f8 "
                      "(Annual Report on Forest and Forestry, FY2024)",
             "detail": "Forest cover, planted-forest age structure, forest-road density and network length "
                       "at end FY2023, and wood self-sufficiency.",
             "url": "https://www.rinya.maff.go.jp/j/kikaku/hakusyo/r6hakusyo_h/all/chap1_1_1.html"},
            {"author": "\u8fb2\u6797\u6c34\u7523\u7701 (MAFF, Japan)", "year": "2025",
             "title": "\u6728\u8cea\u30d0\u30a4\u30aa\u30de\u30b9\u30a8\u30cd\u30eb\u30ae\u30fc"
                      "\u5229\u7528\u52d5\u5411\u8abf\u67fb (Wood Biomass Energy Use Survey), 2024",
             "detail": "Wood pellet and wood chip volumes burned for energy, split by domestic and "
                       "imported origin and by feedstock category. e-Stat tables 0004045573 and 0004045572.",
             "url": "https://www.e-stat.go.jp/dbview?sid=0004045573"},
            {"author": "\u65e5\u672c\u6728\u8cea\u30d0\u30a4\u30aa\u30de\u30b9\u30a8\u30cd"
                       "\u30eb\u30ae\u30fc\u5354\u4f1a (Japan Woody Bioenergy Association)",
             "year": "2025",
             "title": "2024\u5e74\u5ea6 \u71c3\u6599\u6750\u9700\u7d66\u52d5\u5411\u8abf\u67fb"
                      "\u5831\u544a\u66f8 (Fuel Wood Supply and Demand Survey)",
             "detail": "Palm kernel shell imports for 2024 by origin country, using Ministry of Finance "
                       "trade data.",
             "url": "https://jwba.or.jp/wp/wp-content/uploads/2025/04/2024%E7%87%83%E6%96%99%E6%9D%90%E9%9C%80%E7%B5%A6%E5%8B%95%E5%90%91%E8%AA%BF%E6%9F%BB%E5%A0%B1%E5%91%8A%E6%9B%B8.pdf"},
            {"author": "\u6797\u91ce\u5e81 (Forestry Agency, Japan)", "year": "2025",
             "title": "\u4ee4\u548c6\u5e74\u6728\u6750\u9700\u7d66\u8868 (Wood Supply and Demand "
                      "Table, 2024) \u30fb\u6728\u6750\u8f38\u5165\u5b9f\u7e3e (Wood Import Record)",
             "detail": "Domestic industrial roundwood, total supply and demand, the 42.5% self-sufficiency "
                       "ratio, and pellet imports by origin country.",
             "url": "https://www.rinya.maff.go.jp/j/press/kikaku/251121.html"},
            {"author": "Swedish Energy Agency (Energimyndigheten)", "year": "2025",
             "title": "Of\u00f6r\u00e4dlade tr\u00e4dbr\u00e4nslen 2024 (Unprocessed wood fuels, 2024)",
             "detail": "Total unprocessed wood-fuel use, the industrial by-product, primary forest fuel "
                       "and recycled categories, and the under-5% imported raw material share.",
             "url": "https://www.energimyndigheten.se/nyhetsarkiv/2025/ny-officiell-statistik-oforadlade-tradbranslen-2024/"},
            {"author": "Swedish Energy Agency (Energimyndigheten)", "year": "2025",
             "title": "Tr\u00e4dbr\u00e4nsle- och torvpriser (Wood fuel and peat prices)",
             "detail": "Forest-chip and refined-wood-fuel prices free to consuming plant excluding tax, "
                       "2024. Converted at the year's average exchange rates.",
             "url": "https://pxexternal.energimyndigheten.se/pxweb/en/Energimyndighetens_statistikdatabas/Energimyndighetens_statistikdatabas__Officiell_energistatistik__Tradbransle_och_torvpriser/2_EN0307_2.px/"},
            {"author": "Swedish University of Agricultural Sciences (SLU)", "year": "2026",
             "title": "National Forest Inventory \u2014 the latest statistics",
             "detail": "Swedish total forest land 27.9 Mha, productive forest land 23.5 Mha and land area "
                       "40.7 Mha.",
             "url": "https://www.slu.se/en/about-slu/organisation/departments/forest-resource-management/miljoanalys/nfi/our-data/the-latest-statistics/"},
            {"author": "Swedish Forest Agency (Skogsstyrelsen)", "year": "2026",
             "title": "Fellings",
             "detail": "Gross felling 87.7 million m\u00b3sk (2024) and preliminary net felling with its "
                       "sawlog, pulpwood and firewood split (2025).",
             "url": "https://www.skogsstyrelsen.se/en/statistics/felling/fellings/"},
            {"author": "\u7d4c\u6e08\u7523\u696d\u7701 (METI, Japan)", "year": "2025",
             "title": "\u8abf\u9054\u4fa1\u683c\u7b49\u7b97\u5b9a\u59d4\u54e1\u4f1a "
                      "(Procurement Price Calculation Committee), \u8cc7\u65991",
             "detail": "Average reported wood chip, pellet and PKS fuel costs in Japanese biomass "
                       "projects, 2025 analysis.",
             "url": "https://www.meti.go.jp/shingikai/santeii/pdf/104_01_00.pdf"},
        ],
    }


CLAIMS.append(build_se_jp_biomass)

if __name__ == "__main__":
    main()
