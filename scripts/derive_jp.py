"""Derive Japan Sankey flows from METI 総合エネルギー統計 (Comprehensive Energy Statistics).

Source file: stte_<FY>.xlsx, sheet ｴﾈﾙｷﾞｰ単位表（本表） (energy-unit balance table), values in TJ.
Run identically for every fiscal year so datasets stay methodologically consistent.

Sign convention in the METI table: transformation rows are negative for inputs consumed and
positive for outputs produced. Primary supply is row #190000 (国内供給, domestic supply), which
is already net of exports and stock change, so summing the fuel bands reproduces METI's
published headline exactly with no double counting.

"Own use & losses" is taken as the residual of each band (supply - conversion - final - non-energy).
It therefore absorbs transformation losses (coke ovens, refineries, gas works), energy-industry
own use, transmission losses, stock changes and the statistical difference — matching how the
LLNL charts bundle these.
"""
import os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))

# ── fuel bands: chart id -> METI column codes summed into it ───────────────
BANDS = {
    "nuclear": ["$1100"],                  # 原子力発電 (primary heat equivalent)
    "hydro":   ["$0800"],                  # 水力発電（揚水除く）
    "solar":   ["$N111", "$N112"],         # 太陽光発電 + 太陽熱利用
    "wind":    ["$N120"],                  # 風力発電
    "geo":     ["$N160"],                  # 地熱エネルギー
    "bio":     ["$N130"],                  # バイオマスエネルギー
    "wst":     ["$1000"],                  # 未活用エネルギー (waste & recovered)
    "coal":    ["$0100", "$0200"],         # 石炭 + 石炭製品
    "gas":     ["$0500", "$0600"],         # 天然ガス + 都市ガス
    "oil":     ["$0300", "$0400"],         # 原油 + 石油製品
}
BAND_LABEL = {"nuclear": "Nuclear", "hydro": "Hydro", "solar": "Solar", "wind": "Wind",
              "geo": "Geothermal", "bio": "Biomass", "wst": "Waste & recovered energy",
              "coal": "Coal", "gas": "Natural gas & city gas", "oil": "Oil"}

ROW = {
    "supply":   ["#190000"],                        # 国内供給
    "elec":     ["#240000", "#250000"],             # 事業用発電 + 自家用発電
    "heat":     ["#260000", "#270000"],             # 自家用蒸気発生 + 熱供給
    "ind":      ["#610000", "#620000"],             # 農林水産鉱建設業 + 製造業
    "com":      ["#650000"],                        # 業務他
    "res":      ["#700000"],                        # 家庭
    "tpt":      ["#800000"],                        # 運輸
    "ne":       ["#950000"],                        # 非エネルギー利用 (subset of final)
    "ne_ind":   ["#951100", "#951500", "#951700"],  # non-energy within industry
    "ne_com":   ["#951800"],                        # non-energy within commerce
    "ne_res":   ["#952000"],
    "ne_tpt":   ["#953000"],
    "own_use":  ["#301000"],                        # 自家消費
    "td_loss":  ["#305000"],                        # 送配電熱損失
    "statdiff": ["#400000"],                        # 統計誤差
    # product manufacture inside an aggregated band (coke ovens, refineries, gas works,
    # inter-product transfers such as LPG blended into city gas) — must be netted in, or a
    # band like gas double-counts the natural gas that becomes city gas
    "prodmfg":  ["#210000", "#220000", "#230000", "#280000"],
    "stock":    ["#350000"],                        # 転換・消費在庫変動
}
ELEC_COL, HEAT_COL = "$1200", "$1300"


def primary_equivalent_factor(year):
    """METI converts every electricity-generating primary source to primary-energy equivalent
    at one uniform reference efficiency. Recover it from the natural-units sheet, which reports
    the same rows in GWh: factor = generated electricity (GWh x 3.6 TJ) / primary energy (TJ).
    Verified identical for nuclear, hydro, solar PV and wind (41.80% FY2023, 42.41% FY2024)."""
    wb = load_workbook(os.path.join(HERE, f"stte_{year}.xlsx"), read_only=True, data_only=True)
    ws = wb["固有単位表"]
    rows = list(ws.iter_rows(values_only=True))
    col = {str(c).strip(): j for j, c in enumerate(rows[0])
           if c and str(c).strip().startswith("$")}
    ridx = {}
    for i, r in enumerate(rows):
        if r[0] and str(r[0]).strip().startswith("#"):
            ridx.setdefault(str(r[0]).strip(), i)
    gwh = float(rows[ridx["#190000"]][col["$0800"]])       # hydro, in 10^6 kWh
    wb.close()
    val = read_table(year)
    primary = val("#190000", "$0800")
    return (gwh * 3.6) / primary if primary else None


def read_table(year):
    wb = load_workbook(os.path.join(HERE, f"stte_{year}.xlsx"), read_only=True, data_only=True)
    ws = wb["ｴﾈﾙｷﾞｰ単位表（本表）"]
    rows = list(ws.iter_rows(values_only=True))
    col = {str(c).strip(): j for j, c in enumerate(rows[0])
           if c and str(c).strip().startswith("$") and len(str(c).strip()) > 1}
    ridx = {}
    for i, r in enumerate(rows):
        c0 = r[0]
        if c0 and str(c0).strip().startswith("#") and len(str(c0).strip()) > 1:
            ridx.setdefault(str(c0).strip(), i)
    wb.close()

    def val(rowcode, colcode):
        if rowcode not in ridx or colcode not in col:
            return 0.0
        v = rows[ridx[rowcode]][col[colcode]]
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
    return val


def derive(year):
    val = read_table(year)
    flows, audit = [], []

    def add(s, t, v, note=""):
        v = round(v)
        if v > 0:
            flows.append((s, t, v))
            audit.append({"from": s, "to": t, "TJ": v, "derivation": note})

    def cell(rowkey, cols):
        return sum(val(r, c) for r in ROW[rowkey] for c in cols)

    for fid, cols in BANDS.items():
        cs = "+".join(cols)
        # band throughput = domestic supply, net of product manufacture inside the band
        # (so natural gas converted to city gas is not counted twice), plus stock drawdown
        supply = cell("supply", cols) + cell("prodmfg", cols) + cell("stock", cols)
        if supply <= 0:
            continue
        # conversion inputs are negative in the table
        e_in = -cell("elec", cols)
        h_in = -cell("heat", cols)
        add(fid, "elec", e_in, f"-(#240000+#250000)[{cs}] — utility + auto-producer generation")
        add(fid, "heat", h_in, f"-(#260000+#270000)[{cs}] — auto steam + heat supply")
        # final sectors, net of the non-energy use booked inside them
        sec_vals = {}
        for sec, nekey in (("ind", "ne_ind"), ("com", "ne_com"), ("res", "ne_res"), ("tpt", "ne_tpt")):
            v = cell(sec, cols) - cell(nekey, cols)
            sec_vals[sec] = v
            add(fid, sec, v, f"{'+'.join(ROW[sec])}[{cs}] less non-energy {'+'.join(ROW[nekey])}")
        ne = cell("ne", cols)
        add(fid, "ne", ne, f"#950000[{cs}] — non-energy use (feedstocks)")
        own = supply - e_in - h_in - sum(max(0, v) for v in sec_vals.values()) - ne
        add(fid, "own", own,
            f"residual: throughput {supply:,.0f} (#190000 + #210000..#280000 + #350000)[{cs}] "
            f"less conversion, final and non-energy use — energy-industry own use, "
            f"transformation losses, T&D and statistical difference")

    # ── electricity box outputs ────────────────────────────────────────────
    for sec in ("ind", "com", "res", "tpt"):
        add("elec", sec, cell(sec, [ELEC_COL]), f"{'+'.join(ROW[sec])}[{ELEC_COL}]")
    add("elec", "heat", -cell("heat", [ELEC_COL]), f"-(#260000+#270000)[{ELEC_COL}] — electricity into heat supply")
    add("elec", "own", -cell("own_use", [ELEC_COL]) - cell("td_loss", [ELEC_COL]) + cell("statdiff", [ELEC_COL]),
        f"-#301000 - #305000 + #400000 [{ELEC_COL}] — own use, T&D losses, statistical difference")
    ein = sum(f[2] for f in flows if f[1] == "elec")
    add("elec", "rej", ein - sum(f[2] for f in flows if f[0] == "elec"),
        "conversion losses = box inputs - box outputs (balancing item)")

    # ── steam & heat box outputs ───────────────────────────────────────────
    for sec in ("ind", "com", "res"):
        add("heat", sec, cell(sec, [HEAT_COL]), f"{'+'.join(ROW[sec])}[{HEAT_COL}]")
    hin = sum(f[2] for f in flows if f[1] == "heat")
    add("heat", "rej", hin - sum(f[2] for f in flows if f[0] == "heat"),
        "conversion + distribution losses = box inputs - box outputs (balancing item)")

    add("own", "rej", sum(f[2] for f in flows if f[1] == "own"),
        "all energy-industry own use and losses are rejected energy")

    meta = {
        "domestic_primary_supply_TJ": round(val("#190000", "$1400")),
        "final_consumption_TJ": round(val("#500000", "$1400")),
        "electricity_generated_TJ": round(val("#240000", ELEC_COL) + val("#250000", ELEC_COL)),
        "non_energy_use_TJ": round(val("#950000", "$1400")),
        "primary_equivalent_factor": round(primary_equivalent_factor(year), 5),
        # scope of the biomass / waste bands, for the comparability footnote:
        # METI keeps municipal waste (biogenic and fossil alike) out of biomass, and
        # books recovered industrial steam and electricity inside the waste band
        "recovered_heat_elec_TJ": round(val("#190000", "$N250")),
        "refuse_energy_TJ": round(val("#190000", "$N200") - val("#190000", "$N250")),
        "liquid_biofuel_TJ": round(val("#190000", "$N133")),
    }
    return flows, audit, meta


if __name__ == "__main__":
    for y in (2023, 2024):
        fl, au, me = derive(y)
        bands = {}
        for s, t, v in fl:
            if s in BAND_LABEL:
                bands[s] = bands.get(s, 0) + v
        print(f"\n===== JAPAN FY{y} =====   flows: {len(fl)}")
        for k, v in me.items():
            print(f"  {k:34}{v:>13,}")
        print(f"  {'primary supply (sum of bands)':34}{sum(bands.values()):>13,}")
        for k, v in sorted(bands.items(), key=lambda x: -x[1]):
            print(f"     {BAND_LABEL[k]:28}{v:>13,}")
