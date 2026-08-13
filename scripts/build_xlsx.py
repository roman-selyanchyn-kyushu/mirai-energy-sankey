"""Build energy_sankey_data.xlsx — all four datasets, fully auditable.

Sheets:
  README            what the workbook is, sources, how to cite, units
  All flows         tidy long format: one row per flow (pivot-friendly)
  Derivation        every flow with the exact source rows/columns it came from
  Reconciliation    node-by-node balance check + headline totals vs official publications
  SE 2023 .. JP FY2024   one readable matrix per dataset
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = "/Users/romanselyanchyn/Library/CloudStorage/Dropbox/Apps/Projects/Sankey"
DATA = json.load(open(os.path.join(HERE, "datasets.json")))

HEAD_FILL = PatternFill("solid", fgColor="1D232A")
SUB_FILL = PatternFill("solid", fgColor="F3F1EA")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
H2_FONT = Font(bold=True, size=11, color="0F6F8E")
THIN = Side(style="thin", color="D8D5CC")
BORDER = Border(bottom=THIN)
ORDER = ["se2023", "se2024", "jp2023", "jp2024"]
SHEET_NAME = {"se2023": "SE 2023", "se2024": "SE 2024",
              "jp2023": "JP FY2023", "jp2024": "JP FY2024"}


def label_map(d):
    return {n["id"]: n["label"] for n in d["nodes"]}


def style_header(ws, row=1, ncol=8):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, *pairs):
    for col, w in pairs:
        ws.column_dimensions[col].width = w


def sheet_readme(wb):
    ws = wb.create_sheet("README")
    rows = [
        ("Energy flow (Sankey) datasets — Sweden and Japan, 2023 and 2024", "title"),
        ("", ""),
        ("Prepared for the MIRAI Research Group, Kyushu University. Compiled and verified August 2026.", ""),
        ("Interactive charts: https://roman-selyanchyn-kyushu.github.io/mirai-energy-sankey/", ""),
        ("Code and documentation: https://github.com/roman-selyanchyn-kyushu/mirai-energy-sankey", ""),
        ("", ""),
        ("Sources", "h2"),
        ("Sweden", "Eurostat, Complete energy balances (nrg_bal_c) and Production of electricity and derived "
                   "heat by fuel (nrg_bal_peh), calendar years 2023 and 2024, terajoules. Retrieved from the "
                   "Eurostat REST API; dataset update 2 June 2026. Both years final, no provisional flags. "
                   "Eurostat receives these from Energimyndigheten (Swedish Energy Agency)."),
        ("Japan", "METI / Agency for Natural Resources and Energy, Comprehensive Energy Statistics "
                  "(総合エネルギー統計), fiscal years 2023 and 2024 (確報, revised), energy-unit balance "
                  "table, terajoules. Files stte_2023.xlsx and stte_2024.xlsx from enecho.meti.go.jp. "
                  "Japanese fiscal years run April to March."),
        ("", ""),
        ("Units", "h2"),
        ("Native unit", "All source data and the 'value_TJ' column are in terajoules (TJ)."),
        ("Conversions", "1 PJ = 1,000 TJ.   1 TWh = 3,600 TJ.   1 PJ = 0.2778 TWh."),
        ("", ""),
        ("Sheets", "h2"),
        ("All flows", "Tidy long format — one row per flow, all four datasets. Best for pivot tables, R or Python."),
        ("Derivation", "The same flows with the exact source table rows/columns and arithmetic used for each. "
                       "Use this to trace any number back to the official statistics."),
        ("Reconciliation", "Node-by-node balance check (inputs = outputs) plus headline totals compared with "
                           "the official published figures and with the IEA energy Sankey."),
        ("SE 2023 … JP FY2024", "One readable matrix per dataset: rows are sources, columns are destinations."),
        ("", ""),
        ("Method summary", "h2"),
        ("Sweden", "Only additive Eurostat balance components are used. Eurostat's attributed memo aggregates "
                   "(BIOE 'Bioenergy', FE 'Fossil energy') include fuel embodied in delivered heat and would "
                   "double-count, so bioenergy is taken as the residual of RA000 after removing hydro, wind, "
                   "solar, ambient heat and geothermal. CHP fuel inputs are split between the electricity and "
                   "district-heating boxes in proportion to each fuel's gross electricity vs heat output. "
                   "Ambient heat from building heat pumps is allocated 80% residential / 20% commercial."),
        ("Japan", "Aggregated bands (coal + coal products, crude + oil products, natural gas + city gas) net "
                  "out product manufacture inside the band (coke ovens, refineries, gas works, inter-product "
                  "transfers such as the ~96 PJ of LPG blended into city gas), otherwise the natural gas that "
                  "becomes city gas would be counted twice. Stock change is included as supply. Utility and "
                  "auto-producer generation are merged, as are auto-steam boilers and the heat-supply business."),
        ("End use", "Both countries use the LLNL end-use efficiency assumptions to split each sector into "
                    "energy services and rejected energy: residential 65%, commercial 65%, industrial 49%, "
                    "transport 21%. These are assumptions, not measured statistics."),
        ("", ""),
        ("Caution", "h2"),
        ("Cross-country comparison", "Sweden (Eurostat) and Japan (METI) follow different conventions. METI "
                                     "uses gross calorific value and a substitution method for renewable "
                                     "electricity; Eurostat uses net calorific value and counts renewable "
                                     "electricity 1:1. Do not compare the two countries' totals directly "
                                     "without adjusting — see the Reconciliation sheet."),
    ]
    r = 1
    for a, b in rows:
        if b == "title":
            ws.cell(row=r, column=1, value=a).font = TITLE_FONT
        elif b == "h2":
            ws.cell(row=r, column=1, value=a).font = H2_FONT
        else:
            ws.cell(row=r, column=1, value=a).font = Font(bold=bool(b))
            if b:
                c = ws.cell(row=r, column=2, value=b)
                c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    widths(ws, ("A", 24), ("B", 112))
    return ws


def sheet_all_flows(wb):
    ws = wb.create_sheet("All flows")
    ws.append(["country", "year", "year_label", "from_id", "from_label", "to_id", "to_label",
               "value_TJ", "value_PJ", "value_TWh"])
    style_header(ws, ncol=10)
    for k in ORDER:
        d = DATA[k]
        lm = label_map(d)
        for s, t, v in d["flows"]:
            ws.append([d["country"], d["year"], d["yearLabel"], s, lm.get(s, s), t, lm.get(t, t),
                       v, round(v / 1000, 3), round(v / 3600, 3)])
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for c in row:
            c.number_format = "#,##0.000" if c.column > 8 else "#,##0"
    widths(ws, ("A", 10), ("B", 7), ("C", 11), ("D", 10), ("E", 24), ("F", 8), ("G", 24),
           ("H", 13), ("I", 12), ("J", 12))
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    return ws


def sheet_derivation(wb):
    ws = wb.create_sheet("Derivation")
    ws.append(["country", "year_label", "from", "to", "value_TJ",
               "source rows / columns and arithmetic used"])
    style_header(ws, ncol=6)
    for k in ORDER:
        d = DATA[k]
        lm = label_map(d)
        for a in d["audit"]:
            ws.append([d["country"], d["yearLabel"], lm.get(a["from"], a["from"]),
                       lm.get(a["to"], a["to"]), a["TJ"], a["derivation"]])
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.number_format = "#,##0"
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    widths(ws, ("A", 10), ("B", 11), ("C", 24), ("D", 24), ("E", 13), ("F", 96))
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    return ws


def sheet_reconciliation(wb):
    ws = wb.create_sheet("Reconciliation")
    r = 1
    ws.cell(row=r, column=1, value="Reconciliation and balance checks").font = TITLE_FONT
    r += 2

    ws.cell(row=r, column=1, value="1. Node balance — every conversion and sector node must have inputs = outputs").font = H2_FONT
    r += 1
    ws.append(["dataset", "node", "inputs_TJ", "outputs_TJ", "difference_TJ", "status"])
    style_header(ws, row=r, ncol=6)
    hdr = r
    r += 1
    for k in ORDER:
        d = DATA[k]
        lm = label_map(d)
        inn, out = {}, {}
        for s, t, v in d["flows"]:
            out[s] = out.get(s, 0) + v
            inn[t] = inn.get(t, 0) + v
        for nid in ("elec", "heat", "own", "res", "com", "ind", "tpt"):
            if nid not in inn:
                continue
            i, o = inn.get(nid, 0), out.get(nid, 0)
            ws.cell(row=r, column=1, value=d["country"] + " " + d["yearLabel"])
            ws.cell(row=r, column=2, value=lm.get(nid, nid))
            ws.cell(row=r, column=3, value=i).number_format = "#,##0"
            ws.cell(row=r, column=4, value=o).number_format = "#,##0"
            ws.cell(row=r, column=5, value=i - o).number_format = "#,##0"
            ws.cell(row=r, column=6, value="OK" if abs(i - o) <= 2 else "CHECK")
            r += 1
    ws.freeze_panes = None
    r += 1

    ws.cell(row=r, column=1, value="2. Headline totals vs the official publications").font = H2_FONT
    r += 1
    ws.append(["dataset", "quantity", "this workbook (TJ)", "official figure", "source of official figure"])
    style_header(ws, row=r, ncol=5)
    r += 1
    official = {
        "se2023": [("Gross electricity production", 597935, "597,935 TJ = 166.1 TWh", "Eurostat nrg_bal_peh GEP TOTAL"),
                   ("Final energy consumption", 1301927, "1,301,927 TJ", "Eurostat nrg_bal_c, sum of final sectors")],
        "se2024": [("Gross electricity production", 620521, "620,521 TJ = 172.4 TWh", "Eurostat nrg_bal_peh GEP TOTAL"),
                   ("Final energy consumption", 1311379, "1,311,379 TJ", "Eurostat nrg_bal_c, sum of final sectors")],
        "jp2023": [("Domestic primary energy supply", 17557652, "17,558 PJ", "METI press release 25 Apr 2025 (確報)"),
                   ("Final energy consumption", 11509459, "11,509 PJ", "METI balance table row #500000")],
        "jp2024": [("Domestic primary energy supply", 17461344, "17,461 PJ", "METI balance table row #190000"),
                   ("Final energy consumption", 11280435, "11,280 PJ", "METI balance table row #500000")],
    }
    for k in ORDER:
        d = DATA[k]
        for qty, val, off, srctxt in official[k]:
            ws.cell(row=r, column=1, value=d["country"] + " " + d["yearLabel"])
            ws.cell(row=r, column=2, value=qty)
            ws.cell(row=r, column=3, value=val).number_format = "#,##0"
            ws.cell(row=r, column=4, value=off)
            ws.cell(row=r, column=5, value=srctxt)
            r += 1
    r += 1

    ws.cell(row=r, column=1, value="3. Why these charts differ from the IEA energy Sankey").font = H2_FONT
    r += 1
    notes = [
        ("Sweden 2023", "IEA total energy supply 1,892 PJ vs 1,975 PJ here. Per fuel the two agree closely "
                        "(biofuels & waste 573.5 vs 570.4 PJ; hydro 238.3 vs 238.3; wind+solar+other 177.1 vs 177.1). "
                        "Differences: the IEA imputes a fixed 33% nuclear efficiency (529 PJ) where Eurostat reports "
                        "actual reactor heat ~36% (485 PJ), and the IEA headline nets out net electricity exports "
                        "(-103 PJ) which this chart draws explicitly. The IEA's own by-source total is 1,983 PJ, "
                        "within 0.4% of this chart."),
        ("Japan FY2023", "IEA total energy supply 15,844 PJ (calendar 2023) vs METI 17,558 PJ (fiscal 2023). "
                         "The ~1,714 PJ gap decomposes into: gross vs net calorific value ~+970 PJ (gas +354, "
                         "oil +409, coal +204); METI's substitution method for hydro/solar/wind ~+916 PJ (implied "
                         "factor 41.8%), offset by nuclear -193 PJ and geothermal -93 PJ; recovered energy "
                         "(未活用エネルギー) counted only by METI +273 PJ; and fiscal vs calendar year timing "
                         "plus LHV conversion of biomass, about -150 PJ."),
        ("Note", "Because Japanese band widths represent throughput including stock change and intra-band "
                 "transformation, the bands sum to slightly more than METI's domestic-supply headline "
                 "(17,629 vs 17,558 PJ in FY2023, a 0.4% difference). Cite METI's headline figure."),
    ]
    for a, b in notes:
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 58
        r += 1
    widths(ws, ("A", 22), ("B", 78), ("C", 20), ("D", 22), ("E", 40), ("F", 10))
    return ws


def sheet_matrix(wb, key):
    d = DATA[key]
    ws = wb.create_sheet(SHEET_NAME[key])
    lm = label_map(d)
    order = [n["id"] for n in d["nodes"]]
    srcs = [n for n in order if any(s == n for s, _, _ in d["flows"])]
    dsts = [n for n in order if any(t == n for _, t, _ in d["flows"])]
    m = {}
    for s, t, v in d["flows"]:
        m[(s, t)] = m.get((s, t), 0) + v

    ws.cell(row=1, column=1, value=f"{d['title']} — all flows in TJ").font = TITLE_FONT
    ws.cell(row=2, column=1, value=d["source"]).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30
    hdr = 4
    ws.cell(row=hdr, column=1, value="from \\ to")
    for j, t in enumerate(dsts):
        c = ws.cell(row=hdr, column=2 + j, value=lm.get(t, t))
        c.alignment = Alignment(text_rotation=45, horizontal="left")
    ws.cell(row=hdr, column=2 + len(dsts), value="TOTAL OUT")
    for c in range(1, 3 + len(dsts)):
        cell = ws.cell(row=hdr, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for i, s in enumerate(srcs):
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=lm.get(s, s)).font = Font(bold=True)
        tot = 0
        for j, t in enumerate(dsts):
            v = m.get((s, t))
            if v:
                ws.cell(row=r, column=2 + j, value=v).number_format = "#,##0"
                tot += v
        c = ws.cell(row=r, column=2 + len(dsts), value=tot)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
    r = hdr + 1 + len(srcs)
    ws.cell(row=r, column=1, value="TOTAL IN").font = Font(bold=True)
    for j, t in enumerate(dsts):
        tot = sum(v for (s2, t2), v in m.items() if t2 == t)
        c = ws.cell(row=r, column=2 + j, value=tot)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)
    widths(ws, ("A", 26))
    for j in range(len(dsts) + 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_readme(wb)
    sheet_all_flows(wb)
    sheet_derivation(wb)
    sheet_reconciliation(wb)
    for k in ORDER:
        sheet_matrix(wb, k)
    dest = os.path.join(PROJECT, "energy_sankey_data.xlsx")
    wb.save(dest)
    print(f"wrote {dest}  ({os.path.getsize(dest)/1024:.0f} KB)")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
