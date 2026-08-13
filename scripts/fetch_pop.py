"""Pull the population figures the per-capita comparison needs, into sources/population.json.

Sweden comes from SCB's PxWeb API. Japan's Statistics Bureau does not expose an
open API without a key, so its official 人口推計 table is kept in sources/ and
parsed here rather than transcribed.

Both countries are put on a mid-period basis so the denominator matches the
energy year: Sweden the mean of the two year-end stocks bracketing the calendar
year, Japan the 1 October estimate, which is the standard reference date inside
the April-March fiscal year.
"""
import json, os, urllib.request
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCES = os.path.join(os.path.dirname(HERE), "sources")
SCB = ("https://api.scb.se/OV0104/v1/doris/en/ssd/BE/BE0101/BE0101A/BefolkningNy")
JP_FILE = os.path.join(SOURCES, "jp_population_2024_stat_go_jp.xlsx")


def sweden(year):
    """SCB publishes population at 31 December; the mean of the bracketing
    year-ends is the conventional mid-year figure for a calendar-year rate."""
    q = {"query": [{"code": "Region", "selection": {"filter": "item", "values": ["00"]}},
                   {"code": "Alder", "selection": {"filter": "item", "values": ["tot"]}},
                   {"code": "ContentsCode", "selection": {"filter": "item", "values": ["BE0101N1"]}},
                   {"code": "Tid", "selection": {"filter": "item", "values": [str(year-1), str(year)]}}],
         "response": {"format": "json-stat2"}}
    req = urllib.request.Request(SCB, data=json.dumps(q).encode(),
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=90) as r:
        d = json.load(r)
    yrs = list(d["dimension"]["Tid"]["category"]["label"].values())
    assert yrs == [str(year-1), str(year)], yrs
    a, b = d["value"]
    return {"value": round((a+b)/2), "year_end_prev": a, "year_end": b,
            "basis": f"mean of 31 Dec {year-1} and 31 Dec {year}",
            "source": "SCB (Statistics Sweden), table BE0101 Population by region, age and year",
            "url": "https://www.scb.se/en/finding-statistics/statistics-by-subject-area/population/"}


def japan(fy):
    """第1表 of 人口推計: the 総数 row of the 総人口 column, in thousands."""
    wb = load_workbook(JP_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    total = None
    for r in ws.iter_rows(values_only=True):
        label = str(r[0]).replace(" ", "").replace("　", "") if r[0] else ""
        if label == "総数":
            total = next(v for v in r[1:] if isinstance(v, (int, float)))
            break
    wb.close()
    assert total, "総数 row not found"
    return {"value": int(round(total*1000)),
            "basis": f"1 October {fy}, the standard reference date inside fiscal year {fy}",
            "source": "Statistics Bureau of Japan, 人口推計 (Population Estimates), 第1表, 総人口",
            "url": "https://www.stat.go.jp/data/jinsui/2024np/index.html"}


def main():
    out = {"se": {"2024": sweden(2024)}, "jp": {"2024": japan(2024)}}
    path = os.path.join(SOURCES, "population.json")
    with open(path, "w") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)
    print(f"wrote {path}")
    for c, yrs in out.items():
        for y, d in yrs.items():
            print(f"  {c} {y}: {d['value']:>12,}   {d['basis']}")


if __name__ == "__main__":
    main()
